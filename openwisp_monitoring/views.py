import csv
import logging
from collections import OrderedDict
from copy import copy
from datetime import datetime
from io import BytesIO, StringIO

from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string
from pytz import timezone
from pytz import timezone as tz
from pytz.exceptions import UnknownTimeZoneError
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from swapper import load_model
from django.core.serializers import serialize

from .monitoring.exceptions import InvalidChartConfigException

logger = logging.getLogger(__name__)

Chart = load_model("monitoring", "Chart")


class MonitoringApiViewMixin:
    def _get_charts(self, request, *args, **kwargs):
        """Hook to return Chart query."""
        raise NotImplementedError

    def _get_additional_data(request, *args, **kwargs):
        """Hook to return any additonal data that should be included in the response."""
        return {}

    def _validate_custom_date(self, start, end, tmz):
        try:
            start = datetime.strptime(start, "%Y-%m-%d %H:%M:%S")
            end = datetime.strptime(end, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            raise ValidationError(
                "Incorrect custom date format, should be YYYY-MM-DD H:M:S"
            )
        if (end - start).days > 365:
            raise ValidationError("The date range shouldn't be greater than 365 days")
        if start > end:
            raise ValidationError("start_date cannot be greater than end_date")
        now_tz = datetime.now(tz=timezone(tmz)).strftime("%Y-%m-%d %H:%M:%S")
        now = datetime.strptime(now_tz, "%Y-%m-%d %H:%M:%S")
        if start > now:
            raise ValidationError("start_date cannot be greater than today's date")
        if end > now:
            raise ValidationError("end_date cannot be greater than today's date")
        return start, end

    def get(self, request, *args, **kwargs):
        time = request.query_params.get("time", Chart.DEFAULT_TIME)
        start_date = request.query_params.get("start", None)
        end_date = request.query_params.get("end", None)

        # try to read timezone
        tz_name = request.query_params.get("timezone", settings.TIME_ZONE)

        # 🔧 map old / invalid names to valid IANA tz names
        TZ_ALIAS = {
            "Asia/Calcutta": "Asia/Kolkata",
            "asia/Calcutta": "Asia/Kolkata",
            "Asia/kolkata": "Asia/Kolkata",
        }
        tz_name = TZ_ALIAS.get(tz_name, tz_name)

        try:
            tz(tz_name)
        except UnknownTimeZoneError:
            raise ValidationError("Unknown Time Zone")

        # if custom dates are provided then validate custom dates
        if start_date and end_date:
            start_datetime, end_datetime = self._validate_custom_date(
                start_date, end_date, tz_name
            )
            # if valid custom dates then calculate custom days
            time = "1d"
            custom_days = (end_datetime - start_datetime).days
            if custom_days:
                time = f"{custom_days}d"
        if time not in Chart._get_group_map(time).keys():
            raise ValidationError("Time range not supported")
        charts = self._get_charts(request, *args, **kwargs)
        # prepare response data
        data = self._get_charts_data(charts, time, tz_name, start_date, end_date)
        # csv export has a different response
        if request.query_params.get("csv"):
            # Use ``export_format`` instead of ``format`` because DRF reserves
            # the latter for content-negotiation (URL_FORMAT_OVERRIDE) and
            # would 404 on unknown values like "xlsx" / "pdf".
            export_format = (
                request.query_params.get("export_format") or "csv"
            ).lower()
            columns_param = request.query_params.get("columns")
            # Mirror client-side trace synthesis (e.g. ``total`` on traffic
            # charts is computed by chart.js, not returned from the backend)
            # so exports contain the same columns the user sees in the UI.
            data = self._inject_synthesized_traces(data)
            export_data = (
                self._filter_data_by_columns(data, columns_param)
                if columns_param
                else data
            )
            if export_format == "xlsx":
                payload = self._get_xlsx(export_data)
                response = HttpResponse(
                    payload,
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )
                response["Content-Disposition"] = (
                    'attachment; filename="data.xlsx"'
                )
                return response
            if export_format == "pdf":
                payload = self._get_pdf(export_data)
                if payload is None:
                    raise ValidationError(
                        "PDF export requires WeasyPrint to be installed."
                    )
                response = HttpResponse(payload, content_type="application/pdf")
                response["Content-Disposition"] = 'attachment; filename="data.pdf"'
                return response
            response = HttpResponse(
                self._get_csv(export_data), content_type="text/csv"
            )
            response["Content-Disposition"] = "attachment; filename=data.csv"
            return response
        data.update(self._get_additional_data(request, *args, **kwargs))
        return Response(data)

    def _get_chart_additional_query_kwargs(self, chart):
        """Hook to provide additional kwargs to Chart.read."""
        return None

    def _get_charts_data(self, charts, time, timezone, start_date, end_date):
        chart_map = {}
        x_axys = True
        data = OrderedDict({"charts": []})
        for chart in charts:
            # prepare chart dict
            try:
                chart_dict = chart.read(
                    time=time,
                    x_axys=x_axys,
                    timezone=timezone,
                    start_date=start_date,
                    end_date=end_date,
                    additional_query_kwargs=self._get_chart_additional_query_kwargs(
                        chart
                    ),
                )
                if not chart_dict["traces"]:
                    continue
                chart_dict["description"] = chart.description
                chart_dict["title"] = chart.title.format(
                    metric=chart.metric, **chart.metric.tags
                )
                chart_dict["type"] = chart.type
                chart_dict["unit"] = chart.unit
                chart_dict["summary_labels"] = chart.summary_labels
                chart_dict["colors"] = chart.colors
                chart_dict["colorscale"] = chart.colorscale
                for attr in ["fill", "xaxis", "yaxis"]:
                    value = getattr(chart, attr)
                    if value:
                        chart_dict[attr] = value
                if chart.trace_type:
                    chart_dict["trace_type"] = chart.trace_type
                if chart.trace_order:
                    chart_dict["trace_order"] = chart.trace_order
                if chart.calculate_total:
                    chart_dict["calculate_total"] = chart.calculate_total
                if chart.connect_points:
                    chart_dict["connect_points"] = chart.connect_points
                if chart.trace_labels:
                    chart_dict["trace_labels"] = chart.trace_labels
            except InvalidChartConfigException:
                logger.exception(f"Skipped chart for metric {chart.metric}")
                continue
            # get x axys (only once)
            if x_axys and chart_dict["x"] and chart.type != "histogram":
                data["x"] = chart_dict.pop("x")
                x_axys = False
            # prepare to sort the items according to
            # the order in the chart configuration
            key = f'{chart.order} {chart_dict["title"]}'
            chart_map[key] = chart_dict
        # add sorted chart list to chart data
        data["charts"] = list(OrderedDict(sorted(chart_map.items())).values())
        return data

    def _get_csv(self, data):
        header = ["time"]
        columns = [data.get("x")]
        histograms = []
        for chart in data["charts"]:
            if chart["type"] == "histogram":
                histograms.append(chart)
                continue
            for trace in chart["traces"]:
                header.append(self._get_csv_header(chart, trace))
                columns.append(trace[1])
        rows = [header]
        for index, element in enumerate(data.get("x", [])):
            row = []
            for column in columns:
                row.append(column[index])
            rows.append(row)
        for chart in histograms:
            rows.append([])
            rows.append([chart["title"]])
            # Export value as 0 if it is None
            for key, value in chart["summary"].items():
                if chart["summary"][key] is None:
                    chart["summary"][key] = 0
            # Sort Histogram on the basis of value in the descending order
            sorted_charts = sorted(
                chart["summary"].items(), key=lambda x: x[1], reverse=True
            )
            for field, value in sorted_charts:
                rows.append([field, value])
        # write CSV to in-memory file object
        fileobj = StringIO()
        csv.writer(fileobj).writerows(rows)
        return fileobj.getvalue()

    def _get_csv_header(self, chart, trace):
        header = trace[0]
        return f'{header} - {chart["title"]}'

    def _inject_synthesized_traces(self, data):
        """Append the synthetic ``total`` trace on charts that declare
        ``calculate_total=True``. The frontend (chart.js) does this for
        rendering, but the backend response doesn't include it — so the
        export filter would otherwise silently drop a column the user
        ticked in the picker.
        """
        for chart in data.get("charts", []):
            if not chart.get("calculate_total"):
                continue
            traces = chart.get("traces") or []
            if not traces:
                continue
            if any((trace[0] if trace else "") == "total" for trace in traces):
                continue
            length = max((len(t[1]) for t in traces if len(t) > 1), default=0)
            if not length:
                continue
            total = [0] * length
            for trace in traces:
                values = trace[1] if len(trace) > 1 else []
                for j in range(min(length, len(values))):
                    v = values[j]
                    if v is None:
                        continue
                    total[j] += v
            traces.append(["total", total])
        return data

    def _filter_data_by_columns(self, data, columns_param):
        """Return a shallow-copied data dict whose charts/traces only
        include the columns named in columns_param.

        columns_param is a pipe-separated list. For non-histogram charts,
        each item is the CSV header "{trace_name} - {chart_title}". For
        histogram charts, the item is "__hist__:{chart_title}".
        Charts/traces not listed are dropped entirely.
        """
        requested = {c for c in (columns_param or "").split("|") if c}
        if not requested:
            return data
        filtered_charts = []
        for chart in data.get("charts", []):
            if chart.get("type") == "histogram":
                if f"__hist__:{chart.get('title', '')}" in requested:
                    filtered_charts.append(chart)
                continue
            kept_traces = [
                trace
                for trace in chart.get("traces", [])
                if self._get_csv_header(chart, trace) in requested
            ]
            if not kept_traces:
                continue
            chart_copy = copy(chart)
            chart_copy["traces"] = kept_traces
            filtered_charts.append(chart_copy)
        new_data = OrderedDict(data)
        new_data["charts"] = filtered_charts
        return new_data

    def _build_export_rows(self, data):
        """Build (header_row, body_rows, histogram_sections) shared by
        all export formats. histogram_sections is a list of
        (title, [(field, value), ...]).
        """
        header = ["time"]
        columns = [data.get("x") or []]
        histograms = []
        for chart in data.get("charts", []):
            if chart.get("type") == "histogram":
                histograms.append(chart)
                continue
            for trace in chart.get("traces", []):
                header.append(self._get_csv_header(chart, trace))
                columns.append(trace[1])
        body = []
        x_values = data.get("x") or []
        for index in range(len(x_values)):
            row = []
            for column in columns:
                row.append(column[index] if index < len(column) else None)
            body.append(row)
        histogram_sections = []
        for chart in histograms:
            summary = chart.get("summary") or {}
            normalized = {
                key: (0 if value is None else value)
                for key, value in summary.items()
            }
            sorted_items = sorted(
                normalized.items(), key=lambda item: item[1], reverse=True
            )
            histogram_sections.append((chart.get("title", ""), sorted_items))
        return header, body, histogram_sections

    def _get_xlsx(self, data):
        """Build an XLSX workbook of the same data the CSV export contains.

        Time-series traces are written as one sheet; each histogram chart
        gets its own sheet so its key-value layout doesn't collide with
        the time series.
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        header, body, histograms = self._build_export_rows(data)
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Series"
        if header:
            ws.append(header)
            for col_idx, value in enumerate(header, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(
                    14, min(40, len(str(value)) + 2)
                )
        for row in body:
            ws.append(row)
        for title, items in histograms:
            sheet_title = (title or "Histogram")[:31] or "Histogram"
            base = sheet_title
            counter = 2
            while sheet_title in wb.sheetnames:
                suffix = f" ({counter})"
                sheet_title = base[: 31 - len(suffix)] + suffix
                counter += 1
            hist_ws = wb.create_sheet(title=sheet_title)
            hist_ws.append([title])
            hist_ws.append(["Field", "Value"])
            for field, value in items:
                hist_ws.append([field, value])
            hist_ws.column_dimensions["A"].width = 32
            hist_ws.column_dimensions["B"].width = 18
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _get_pdf(self, data):
        """Render the selected data as a paginated PDF table.

        Returns None if WeasyPrint is unavailable so the caller can
        surface a clean error.
        """
        try:
            from weasyprint import HTML
        except ImportError:
            logger.exception("WeasyPrint not available; cannot render PDF export.")
            return None
        header, body, histograms = self._build_export_rows(data)
        # When called from the device-scoped API, the subclass sets
        # ``self.instance`` to the Device before super().get() runs, so we
        # can title the PDF with the device name. Falls back to a generic
        # heading for dashboard / multi-tenant exports.
        device = getattr(self, "instance", None)
        title = getattr(device, "name", None) or "Timeseries Export"
        html = render_to_string(
            "monitoring/timeseries_export_pdf.html",
            {
                "title": title,
                "header": header,
                "rows": body,
                "histograms": histograms,
                "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
        )
        return HTML(string=html).write_pdf()

from django.shortcuts import render
 
# def map_view(request):
#     return render(request, "admin/dashboard/map.html")

from django.shortcuts import render
from django.http import JsonResponse

# def map_view(request):
#     return render(request, "map.html")

# views.py
from django.http import JsonResponse
#from openwisp_controller.config.models import Device  # Adjust if your import path differs
Device = load_model("config", "Device")

def devices_geojson(request):
    features = []
    qs = Device.objects.exclude(lat=None).exclude(lon=None)[:500]

    for d in qs:
        try:
            lon = float(d.lon)
            lat = float(d.lat)
        except (TypeError, ValueError):
            continue  # skip bad coords

        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [lon, lat]},  # GeoJSON is [lon, lat]
            "properties": {
                "name": d.name,
                "status": d.get_health_status_display(),  # 'ok', 'critical', etc.
                "id": str(d.id),
            },
        })

    return JsonResponse({"type": "FeatureCollection", "features": features})

def map_view(request):
    devices = Device.objects.all()
    geojson_data = serialize(
        'geojson',
        devices,
        geometry_field='location',  # Make sure this matches your PointField name
        fields=('name', 'ip_address')
    )
    return render(request, 'admin/dashboard/map.html', {'geojson_data': geojson_data})